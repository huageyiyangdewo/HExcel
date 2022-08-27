import os
import re
import logging


from openpyxl import load_workbook

# 允许的判断运算符
SAFE_JUDGE_OPERATOR = ["==", ">", "<", ">=", "<="]
# 允许的算术运算符
SAFE_LOGIC_OPERATOR = ["=", "+", "-", "*", "/"]


class ParseConf(object):

    def __init__(self, file_path):
        self.file_path = file_path
        self.sheet = load_workbook(file_path)["conf"]

    def handle_simple_condition(self, condition):

        row = condition[0].strip().strip("[").strip("]")
        column = condition[1].strip().strip("[").strip("]")

        if len(row.split("|")) < 2:
            return False
        if len(column.split("|")) < 2:
            return False

        if len(condition) > 2:
            # 有判断运算符
            operate = condition[2].strip()
            if operate == "=":
                operate = "=="
            if operate not in SAFE_JUDGE_OPERATOR:
                return False
            value = condition[3].strip()

        else:
            # 无判断运算符
            operate = None
            value = None

        condition_cell = {
            "position": {
                "row": row,
                "column": column,
            },
            "operator": operate,
            "value": value
        }
        return condition_cell

    def handle_condition(self, condition_text, split_operate):
        condition_list = []

        if split_operate is None:
            # 单个条件
            # ([3|导体,4|补偿率|1],>=,1.37)
            condition = condition_text.strip("(").strip(")").split(",")

            result = self.handle_simple_condition(condition)
            if result is False:
                return False
            condition_list.append(result)
        else:
            # 多条件定位
            # ([3|导体,4|补偿率|1],>=,1.37)or([3|XLPE绝缘,4|补偿率],<=,23)

            condition = condition_text.split(split_operate)
            condition = [c.strip("(").strip(")").split(",") for c in condition]

            for c in condition:
                result = self.handle_simple_condition(c)
                if result is False:
                    return False
                condition_list.append(result)

        return condition_list

    def handle_modify_range(self, modify_range):
        if modify_range != "全部":
            result = modify_range.split(",")
        else:
            result = ["all"]

        return result

    def handle_modify(self, condition_text, split_operate):
        condition_list = []

        if split_operate is None:
            # ([3|导体,4|补偿率|1],=,11)
            condition = condition_text.strip("(").strip(")").split(",")
            result = self.handle_simple_modify(condition)
            if result is False:
                return False
            condition_list.append(result)
        else:
            # 多条件定位
            # ([3|导体,4|补偿率|1],=,11)and([3|导体,4|补偿率|1],=,11)
            condition = condition_text.split(split_operate)
            condition = [c.strip("(").strip(")").split(",") for c in condition]

            for c in condition:
                result = self.handle_simple_modify(c)
                if result is False:
                    return False
                condition_list.append(result)

        return condition_list

    def handle_simple_modify(self, condition):

        row = condition[0].strip().strip("[").strip("]")
        column = condition[1].strip().strip("[").strip("]")

        if len(row.split("|")) < 2:
            return False
        if len(column.split("|")) < 2:
            return False

        if len(condition) != 4:
            # 格式不符合
            return False

        else:
            operate = condition[2].strip()

            if operate not in SAFE_LOGIC_OPERATOR:
                return False
            value = condition[3].strip()

        modify_cell = {
            "position": {
                "row": row,
                "column": column,
            },
            "operator": operate,
            "value": value
        }
        return modify_cell

    def parse(self):

        conf = dict()
        """
            {
        "YJV": {
            "sheet_name": "等芯",
            "condition_cell": [{
                'position': {
					'row': '3|导体',
					'column': '4|补偿率|1'
				},
                "operator": ">",
                "value": 10
            }],
            "modify_cell": [
                {
                    'position': {
					'row': '3|导体',
					'column': '4|补偿率|1'
				    },
				    "operator": "=",
                    "value": 10
                }
            ],
            "range": "全部"
        }
    }
        """
        max_row = self.sheet.max_row
        for i in range(2, max_row + 1):
            file_type = self.sheet.cell(row=i, column=1).value
            sheet_name = self.sheet.cell(row=i, column=2).value
            condition_text = self.sheet.cell(row=i, column=3).value
            modify_text = self.sheet.cell(row=i, column=4).value
            modify_range = self.sheet.cell(row=i, column=5).value

            if not all([file_type, sheet_name, condition_text, modify_text, modify_range]):
                return False, "第 %d 行,配置文件格式有空值,请检查!" % i

            file_type = str(file_type).strip()
            sheet_name = str(sheet_name).strip()
            condition_text = str(condition_text).strip()
            modify_text = str(modify_text).strip()
            modify_range = str(modify_range).strip()

            if condition_text.find(")and(") != -1:
                split_operate = ")and("
                condition_relationship = "and"
                condition_list = self.handle_condition(condition_text, split_operate)
                if condition_list is False:
                    return False, "第 %d 行,配置文件格式错误,请检查!" % i
            elif condition_text.find(")or(") != -1:
                split_operate = ")or("
                condition_list = self.handle_condition(condition_text, split_operate)
                condition_relationship = "or"
                if condition_list is False:
                    return False, "第 %d 行,配置文件格式错误,请检查!" % i
            else:
                split_operate = None
                condition_list = self.handle_condition(condition_text, split_operate)
                condition_relationship = None
                if condition_list is False:
                    return False, "第 %d 行,配置文件格式错误,请检查!" % i

            condition_cell = []
            condition_cell.extend(condition_list)
            range_list = self.handle_modify_range(modify_range)

            if modify_text.find(")and(") != -1:
                split_operate = ")and("
                modify_list = self.handle_modify(modify_text, split_operate)
                if modify_list is False:
                    return False, "第 %d 行,配置文件格式错误,请检查!" % i
            elif modify_text.find(")or(") != -1:
                return False, "第 %d 行,配置文件格式错误,请检查!" % i
            else:
                split_operate = None
                modify_list = self.handle_modify(modify_text, split_operate)
                if modify_list is False:
                    return False, "第 %d 行,配置文件格式错误,请检查!" % i

            modify_cell = []
            modify_cell.extend(modify_list)

            cell = {
                "condition_cell": condition_cell,
                "condition_relationship": condition_relationship,
                "modify_cell": modify_cell,
                "range": range_list
            }

            if conf.get(file_type):
                sheet_conf = conf[file_type]
                if sheet_conf.get(sheet_name):
                    sheet_conf[sheet_name].append(cell)

                else:
                    sheet_conf[sheet_name] = [cell]
            else:
                sheet_conf = {sheet_name: [cell]}
                conf[file_type] = sheet_conf

        return True, conf


def find_file_types(file_name, file_types):
    file_name = file_name.split('.')[0]
    file_type = file_name

    for f in file_types:
        if file_name.startswith(f):
            file_type = f
            break
    return file_type


def get_row_column(position):
    if len(position["column"]) >= 3:
        return [0, 0]
    row = int(position['row'])
    column = [x.upper() for x in list(position['column'])]

    column_num = 0
    index = 0
    for i in range(len(column)-1, -1, -1):

        if i == len(column)-1:
            num = ord(column[i]) - 64
        else:
            num = index * 26 * (ord(column[i]) - 64)
        column_num += num
        index += 1

    return [row, column_num]


def modify_cell_value_simple(sheet, modify_cell):
    max_row = sheet.max_row

    for c in modify_cell:
        row, column = get_cell_real_position(sheet, c)
        if row is None or column is None:
            continue
        modify_operator = c['operator']
        modify_value = c['value']
        for r_num in range(row + 1, max_row+1):

            modify_cell_value(sheet, r_num, column, modify_operator, modify_value)


def modify_cell_value_with_operate(sheet, sheet_data_only, modify_cell, condition_cell):
    max_row = sheet.max_row

    c_row, c_column = get_cell_real_position(sheet, condition_cell)
    if c_row is None or c_column is None:
        return
    operator = condition_cell['operator']
    value = condition_cell["value"]

    for row in range(c_row+1, max_row+1):
        cell_value = sheet_data_only.cell(row=row, column=c_column).value
        if cell_value is None:
            continue
        cell_value = str(cell_value)

        judge = cell_value + operator + value
        if eval(judge):

            for c in modify_cell:

                _, m_column = get_cell_real_position(sheet, c)

                modify_operator = c['operator']
                modify_value = c['value']
                modify_cell_value(sheet, row, m_column, modify_operator, modify_value)


def modify_cell_value(sheet, row, column, operator, value):
    """
    修改单元格的值
    :param sheet:
    :param row:
    :param column:
    :param operator:
    :param value:
    :return:
    """
    if sheet.cell(row=row, column=column).value is None:
        return

    if operator == "=":
        sheet.cell(row=row, column=column).value = value
    else:
        cell_value = sheet.cell(row=row, column=column).value
        result = calc_modify_result(str(cell_value), operator, value)
        sheet.cell(row=row, column=column).value = result


def create_directory(path):
    if not os.path.exists(path):
        os.mkdir(path)


def judge_translate_operator(operator):
    safe_operator = ["=", ">", "<", ">=", "<="]
    if operator not in safe_operator:
        return False
    else:
        if operator == "=":
            operator = "=="

        return operator


def get_cell_real_position(sheet, position):
    max_column = sheet.max_column

    row = position['position']['row']
    column = position['position']['column']
    if len(column.split("|")) == 3:
        column_num, column_name, column_label = column.split("|")
        column_label = int(column_label)
    else:
        column_num, column_name = column.split("|")
        column_label = None

    row_num, row_name = row.split("|")
    row_num = int(row_num)
    column_num = int(column_num)

    real_row_num = None
    real_column_num = None

    temp_column = 0  # 合并单元格中最大的列数 + 1
    temp_row = 0
    flag = False
    for r in range(1, max_column):
        cell = sheet.cell(row=row_num, column=r)
        c_value = cell.value
        if c_value == row_name:
            flag = True
            # 记录此时的列的数，需要列号时，从这个数开始循环找
            temp_row = r
        else:
            if flag:
                if c_value is not None:

                    temp_column = r
                    break

    count = 1
    for c in range(temp_row, temp_column + 1):
        cell = sheet.cell(row=column_num, column=c)
        c_value = cell.value
        if c_value == column_name:
            if column_label is None:
                real_row_num = column_num
                real_column_num = c
                break
            else:
                # 需要根据 column_label 来定位
                if count == column_label:
                    real_row_num = column_num
                    real_column_num = c
                    break
                count += 1

    return real_row_num, real_column_num


def calc_modify_result(cell_value, operator, value):
    r = cell_value + operator + value

    r = eval(r)

    if type(r) is int:
        pass
    elif type(r) is float:
        r = "%0.2f" % r

    return r


if __name__ == "__main__":
    l = ParseConf("配置文件111.xlsx")
    c = l.parse()
    print(c)

    # l = create_logging(os.getcwd())
    # l.info("xxx")
    #
    # p = {"row": 2, "column": "aa"}
    # c = get_row_column(p)
    # print(c)


