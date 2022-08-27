import os
import re
import logging
import time
from math import ceil
from concurrent.futures import ThreadPoolExecutor
from queue import Queue

QUEUE = Queue()

from openpyxl import load_workbook

from utils import find_file_types, modify_cell_value, create_directory, \
    modify_cell_value_simple, modify_cell_value_with_operate, \
    calc_modify_result, get_cell_real_position


class HandleExcel(object):

    def __init__(self, conf, directory_path):
        self.conf = conf
        self.directory_path = directory_path
        self.files_list = None

    def find_all_excel_file(self):
        files = os.listdir(self.directory_path)
        result = []
        for f in files:
            file_type = find_file_types(f, self.conf.keys())
            if self.conf.get(file_type):
                # 是需要修改的文件
                result.append(f)

        self.files_list = filter(lambda x: re.findall("(.xlsx)$", x), result)
        # self.files_list = filter(lambda x: re.findall("(.xlsx)|(.xls)$", x), files)
        # self.files_list = map(lambda x: self.directory_path + "\\" + x, files_list)

    def handle(self, progress_bar):

        self.find_all_excel_file()
        a = time.time()

        with ThreadPoolExecutor(5) as threadPool:

            count = 0
            for file_name in self.files_list:
                count += 1

                threadPool.submit(calc, self.directory_path, self.conf, file_name)

            num = 0
            while True:
                QUEUE.get()
                time.sleep(0.3)
                num += 1
                progress_bar.setValue(ceil((num / count) * 100))
                if num == count:
                    break

            print("count", count)
        print(time.time()-a)


def calc(directory_path, conf, file_name):
    file_type = find_file_types(file_name, conf.keys())
    sheet_conf = conf[file_type]

    # 打开两个，一个用来判断，一个用来修改
    wb_data_only = load_workbook(directory_path + "\\" + file_name, data_only=True)
    wb = load_workbook(directory_path + "\\" + file_name)

    is_save = True
    for sheet_name, values in sheet_conf.items():
        sheet = wb[sheet_name]
        sheet_data_only = wb_data_only[sheet_name]
        for cell_conf in values:
            modify_range = cell_conf['range']
            if modify_range[0] != "all":
                if file_name not in modify_range:
                    # 不在修改文件范围内
                    is_save = False
                    break

            if cell_conf["condition_relationship"] is None:
                # 条件只有一个
                condition_cell = cell_conf["condition_cell"][0]
                operator = condition_cell['operator']
                if operator is None:
                    temp_row, temp_column = get_cell_real_position(sheet, condition_cell)
                    if not all([temp_row, temp_column]):
                        # 找不到判断的单元格, 不修改
                        continue

                    # 无操作符,所有的都要修改
                    modify_cell = cell_conf['modify_cell']
                    modify_cell_value_simple(sheet, modify_cell)

                else:
                    # 有 条件操作符
                    modify_cell = cell_conf['modify_cell']
                    modify_cell_value_with_operate(sheet, sheet_data_only, modify_cell, condition_cell)

            else:
                # 条件有多个, 需要先遍历行
                # 判断结果
                condition_relationship = str(cell_conf['condition_relationship']).lower()
                if condition_relationship == "and":

                    max_row = sheet.max_row
                    condition_cells = cell_conf['condition_cell']
                    # 这里假定 每个条件的其实行数必须要相同
                    start_num, start_column = get_cell_real_position(sheet, condition_cells[0])
                    if start_num is None or start_column is None:
                        # 说明配置文件有问题，找不到具体的 cell
                        continue

                    for row in range(start_num + 1, max_row + 1):  # 遍历行
                        result = True
                        for c_conf in condition_cells:  # 遍历判断条件
                            c_row, c_column = get_cell_real_position(sheet, c_conf)
                            if c_row is None or c_column is None:
                                # 判断的列不存在
                                result = False
                                break

                            operator = c_conf['operator']
                            if operator is None:
                                # 无操作符，所有条件都满足
                                continue
                            else:

                                cell_value = sheet_data_only.cell(row=row, column=c_column).value
                                if cell_value is None:
                                    result = False
                                    # 有一个不满足，就不能修改, 这里要注意的是，这里是空值
                                    break
                                cell_value = str(cell_value)
                                value = c_conf["value"]

                                judge = cell_value + operator + value
                                if not eval(judge):
                                    # 不满足
                                    result = False
                                    break

                        if result:
                            # 所有条件都满足

                            modify_cell = cell_conf['modify_cell']

                            for c in modify_cell:
                                _, column = get_cell_real_position(sheet, c)
                                modify_operator = c['operator']
                                modify_value = c['value']
                                modify_cell_value(sheet, row, column, modify_operator, modify_value)

                elif condition_relationship == "or":

                    max_row = sheet.max_row
                    condition_cells = cell_conf['condition_cell']
                    # 这里假定 每个条件的其实行数必须要相同
                    start_num, start_column = get_cell_real_position(sheet, condition_cells[0])
                    if start_num is None or start_column is None:
                        # 说明配置文件有问题，找不到具体的 cell
                        continue

                    for row in range(start_num + 1, max_row + 1):  # 遍历行
                        result = False
                        for c_conf in condition_cells:  # 遍历判断条件
                            c_row, c_column = get_cell_real_position(sheet, c_conf)
                            if not all([c_row, c_column]):
                                # 判断的单元格不存在
                                continue

                            operator = c_conf['operator']
                            if operator is None:
                                # 无操作符，所有条件都满足
                                result = True
                                break
                            else:
                                cell_value = sheet_data_only.cell(row=row, column=c_column).value
                                if cell_value is None:
                                    # 不满足，继续判断
                                    continue
                                cell_value = str(cell_value)
                                value = c_conf["value"]

                                judge = cell_value + operator + value
                                if eval(judge):
                                    # 满足
                                    result = True
                                    break

                        if result:
                            # 有条件都满足

                            modify_cell = cell_conf['modify_cell']

                            for c in modify_cell:
                                _, column = get_cell_real_position(sheet, c)

                                modify_operator = c['operator']
                                value = c['value']

                                modify_cell_value(sheet, row, column, modify_operator, value)

                else:
                    logging.error("暂时不支持的逻辑运算符！")
                    break

    if is_save:
        direc_path = directory_path + "\\" + "处理后的文件"
        create_directory(direc_path)
        path = direc_path + "\\" + file_name
        wb.save(path)

    wb.close()
    wb_data_only.close()
    QUEUE.put("1")


def set_process_bar(process_bar, count):
    num = 0
    while QUEUE.get():

        num += 1
        process_bar.setValue(ceil((num / count) * 100))
        if num == count:
            break


if __name__ == "__main__":
    c = {'YJV': {'等芯': [{'condition_cell': [{'position': {'row': '5', 'column': 'x'}, 'operator': '大于', 'value': '10'}],
                         'condition_relationship': None,
                         'modify_cell': [{'position': {'row': '5', 'column': 'x'}, 'value': '10'},
                                         {'position': {'row': '5', 'column': 'v'}, 'value': '1.75'}], 'range': []},
                        {'condition_cell': [{'position': {'row': '5', 'column': 'T'}, 'operator': None, 'value': None}],
                         'condition_relationship': None,
                         'modify_cell': [{'position': {'row': '5', 'column': 't'}, 'value': '10'}], 'range': []}],
                 '不等芯': [
                     {'condition_cell': [{'position': {'row': '5', 'column': 'o'}, 'operator': '大于', 'value': '10'}],
                      'condition_relationship': None,
                      'modify_cell': [{'position': {'row': '5', 'column': 'o'}, 'value': '10'}], 'range': []}],
                 'Sheet1': [
                     {'condition_cell': [{'position': {'row': '5', 'column': 'b'}, 'operator': '大于', 'value': '10'}],
                      'condition_relationship': None,
                      'modify_cell': [{'position': {'row': '5', 'column': 'b'}, 'value': '10'}], 'range': []}]}}

    h = HandleExcel(c, 'E:\\learn_pandas\\HandleExcel')
    h.find_all_excel_file()
    a = list(h.files_list)
    print(a[0])
    print(os.path.split(a[0]))